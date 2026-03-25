from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_FILE = "Estimation_Par_Classe.xlsx"

sections = [
    {
        "classe": "Utilisateur",
        "responsable": "Hammouchi Louay",
        "rows": [
            ("Inscription", 1.0),
            ("Connexion", 0.5),
            ("Déconnexion", 0.5),
            ("Modification des données", 1.5),
            ("Suppression de compte", 0.75),
            ("Mot de passe oublié", 0.5),
            ("Gestion rôle administrateur", 0.5),
        ],
    },
    {
        "classe": "UserDAO",
        "responsable": "Hammouchi Louay",
        "rows": [
            ("createUser", 0.75),
            ("findByEmail", 0.75),
            ("findByPhoneNumber", 0.2),
            ("updateUsername", 0.75),
            ("updateEmail", 0.2),
            ("updatePassword", 0.2),
            ("updatePhoneNumber", 0.2),
            ("deleteUser", 0.75),
            ("emailExists", 0.5),
            ("phoneNumberexists", 0.15),
        ],
    },
    {
        "classe": "Authentification",
        "responsable": "Hammouchi Louay",
        "rows": [
            ("registerUser", 0.5),
            ("loginByEmail", 0.25),
            ("loginByPhoneNumber", 0.1),
            ("logout", 0.25),
            ("modifyUsername", 0.5),
            ("modifyEmail", 0.15),
            ("modifyPassword", 0.15),
            ("modifyPhoneNumber", 0.15),
            ("deleteAccount", 0.5),
            ("forgotPasswordByEmail", 0.5),
            ("forgotPasswordByPhone", 0.25),
            ("resetPassword", 0.5),
        ],
    },
    {
        "classe": "BaseDonnees",
        "responsable": "Hammouchi Louay",
        "rows": [
            ("getConnection", 0.35),
            ("configurationJDBC", 0.25),
            ("GestionErreursConnexion", 0.1),
        ],
    },
]

wb = Workbook()
ws = wb.active
ws.title = "Estimation"

headers = [
    "Classe",
    "Méthode",
    "Estimation (h)",
    "Responsable",
    "Temps effectif de réalisation (h)",
]

header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(color="FFFFFF", bold=True)
section_fill = PatternFill("solid", fgColor="DCE6F1")
total_fill = PatternFill("solid", fgColor="E2F0D9")
center = Alignment(horizontal="center", vertical="center")
left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

current_row = 1
for section in sections:
    # Section header
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    start_data_row = current_row + 1

    # Section data
    for i, (method, estimate) in enumerate(section["rows"]):
        row = start_data_row + i
        ws.cell(row=row, column=1, value=section["classe"] if i == 0 else "")
        ws.cell(row=row, column=2, value=method)
        ws.cell(row=row, column=3, value=estimate)
        ws.cell(row=row, column=4, value=section["responsable"] if i == 0 else "")
        ws.cell(row=row, column=5, value="")

        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            if col in (1, 3, 4, 5):
                cell.alignment = center
            else:
                cell.alignment = left_wrap

        ws.cell(row=row, column=3).number_format = "0.00"
        if i == 0:
            ws.cell(row=row, column=1).fill = section_fill

    # Total row
    total_row = start_data_row + len(section["rows"])
    ws.cell(row=total_row, column=1, value="Totale")
    ws.cell(row=total_row, column=2, value="")
    ws.cell(row=total_row, column=3, value=f"=SUM(C{start_data_row}:C{total_row-1})")
    ws.cell(row=total_row, column=4, value="")
    ws.cell(row=total_row, column=5, value="")

    for col in range(1, 6):
        cell = ws.cell(row=total_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = total_fill
        cell.border = border
        cell.alignment = center if col != 2 else left_wrap

    ws.cell(row=total_row, column=3).number_format = "0.00"

    # Empty row between sections
    current_row = total_row + 2

# Column widths
widths = {1: 20, 2: 34, 3: 18, 4: 24, 5: 34}
for col, width in widths.items():
    ws.column_dimensions[get_column_letter(col)].width = width

ws.freeze_panes = "A2"
wb.save(OUTPUT_FILE)
print(f"Generated {OUTPUT_FILE}")
