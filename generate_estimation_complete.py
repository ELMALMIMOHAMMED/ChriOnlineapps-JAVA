from pathlib import Path
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_FILE = "Estimation_Par_Classe_Complet.xlsx"
RESPONSABLE = "Hammouchi Louay"

CLASS_PATTERN = re.compile(r"\bclass\s+([A-Za-z_][A-Za-z0-9_]*)")
METHOD_PATTERN = re.compile(
    r"^\s*(?:public|protected|private)\s+"
    r"(?:static\s+)?(?:final\s+)?(?:synchronized\s+)?"
    r"(?:<[^>]+>\s*)?[\w\[\]<>?,\s]+\s+([A-Za-z_][A-Za-z0-9_]*)\s*"
    r"\([^;{}]*\)\s*(?:throws\s+[^{]+)?\{"
)
CTOR_PATTERN = re.compile(
    r"^\s*(?:public|protected|private)\s+([A-Za-z_][A-Za-z0-9_]*)\s*"
    r"\([^;{}]*\)\s*(?:throws\s+[^{]+)?\{"
)


def estimate_hours(method_name: str) -> float:
    name = method_name.lower()
    if name.startswith(("get", "set", "is")):
        return 0.2
    if name in {"main", "run"}:
        return 0.5
    if "test" in name:
        return 0.25
    if any(k in name for k in ["find", "exists", "login", "logout", "parse", "map", "json"]):
        return 0.5
    if any(k in name for k in ["create", "register", "add", "remove", "delete", "update", "modify"]):
        return 0.75
    if any(k in name for k in ["checkout", "handle", "route", "process", "connect", "commande", "cart"]):
        return 1.0
    return 0.5


def parse_java_file(file_path: Path):
    text = file_path.read_text(encoding="utf-8", errors="ignore")
    class_match = CLASS_PATTERN.search(text)
    if not class_match:
        return None, []

    class_name = class_match.group(1)
    methods = []

    for line in text.splitlines():
        m = METHOD_PATTERN.match(line)
        if m:
            method_name = m.group(1)
            if method_name not in methods:
                methods.append(method_name)
            continue

        c = CTOR_PATTERN.match(line)
        if c and c.group(1) == class_name:
            ctor_name = f"{class_name} (constructeur)"
            if ctor_name not in methods:
                methods.append(ctor_name)

    return class_name, methods


def collect_sections(root: Path):
    sections = []
    java_files = sorted(root.rglob("*.java"))

    for java_file in java_files:
        # Ignore generated/python helper scripts folders if any Java appears there.
        if any(part.startswith(".") for part in java_file.parts):
            continue

        class_name, methods = parse_java_file(java_file)
        if not class_name:
            continue

        row_data = []
        if methods:
            for method in methods:
                row_data.append((method, estimate_hours(method)))
        else:
            row_data.append(("(aucune méthode détectée)", 0.0))

        sections.append({
            "classe": class_name,
            "responsable": RESPONSABLE,
            "rows": row_data,
            "source": str(java_file.relative_to(root)).replace("\\", "/"),
        })

    return sections


def build_workbook(sections):
    wb = Workbook()
    ws = wb.active
    ws.title = "Estimation Complete"

    headers = [
        "Classe",
        "Méthode",
        "Estimation (h)",
        "Responsable",
        "Temps effectif de réalisation (h)",
    ]

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    section_fill = PatternFill("solid", fgColor="DCE6F1")
    total_fill = PatternFill("solid", fgColor="E2F0D9")
    global_fill = PatternFill("solid", fgColor="FFD966")
    center = Alignment(horizontal="center", vertical="center")
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    current_row = 1
    total_cells = []

    for section in sections:
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=c, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        start_data_row = current_row + 1

        for i, (method, estimate) in enumerate(section["rows"]):
            row = start_data_row + i
            ws.cell(row=row, column=1, value=section["classe"] if i == 0 else "")
            ws.cell(row=row, column=2, value=method)
            ws.cell(row=row, column=3, value=estimate)
            ws.cell(row=row, column=4, value=section["responsable"] if i == 0 else "")
            ws.cell(row=row, column=5, value="")

            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                if col in (1, 3, 4, 5):
                    cell.alignment = center
                else:
                    cell.alignment = left_wrap

            ws.cell(row=row, column=3).number_format = "0.00"
            if i == 0:
                ws.cell(row=row, column=1).fill = section_fill
                ws.cell(row=row, column=1).comment = None
                ws.cell(row=row, column=2).value = f"{method}  [source: {section['source']}]"

        total_row = start_data_row + len(section["rows"])
        ws.cell(row=total_row, column=1, value="Totale")
        ws.cell(row=total_row, column=2, value="")
        ws.cell(row=total_row, column=3, value=f"=SUM(C{start_data_row}:C{total_row-1})")
        ws.cell(row=total_row, column=4, value="")
        ws.cell(row=total_row, column=5, value="")

        total_cells.append(f"C{total_row}")

        for col in range(1, 6):
            cell = ws.cell(row=total_row, column=col)
            cell.font = Font(bold=True)
            cell.fill = total_fill
            cell.border = border
            cell.alignment = center if col != 2 else left_wrap

        ws.cell(row=total_row, column=3).number_format = "0.00"
        current_row = total_row + 2

    # Global total at bottom
    ws.cell(row=current_row, column=1, value="TOTAL GLOBAL")
    ws.cell(row=current_row, column=3, value=f"=SUM({','.join(total_cells)})" if total_cells else 0)
    for col in range(1, 6):
        cell = ws.cell(row=current_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = global_fill
        cell.border = border
        cell.alignment = center
    ws.cell(row=current_row, column=3).number_format = "0.00"

    widths = {1: 24, 2: 52, 3: 18, 4: 24, 5: 34}
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"
    return wb


def main():
    root = Path(__file__).resolve().parent
    sections = collect_sections(root)

    # Keep deterministic order by class name
    sections.sort(key=lambda s: s["classe"].lower())

    wb = build_workbook(sections)
    wb.save(root / OUTPUT_FILE)
    print(f"Generated {OUTPUT_FILE} with {len(sections)} classes")


if __name__ == "__main__":
    main()
