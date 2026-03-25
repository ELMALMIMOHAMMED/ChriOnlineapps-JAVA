from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_FILE = "Cahier_Charges_Produits.xlsx"

headers = [
    "Task ID", "Module", "Tâche", "Type", "Description", "Effort (h)",
    "Durée (j)", "Dépend de", "Priorité", "Risque", "Responsable", "Statut", "Notes"
]

rows = [
    ["AUTH-1", "Authentification", "Concevoir schema users", "Analysis", "Table users avec contraintes unicité email et index", 4, None, "", "P1", "Moyen", "Backend Lead", "Non démarré", "Base pour auth"],
    ["AUTH-2", "Authentification", "Implémenter hash mot de passe", "Dev", "Remplacer stockage clair par hash sécurisé", 3, None, "AUTH-1", "P0", "Élevé", "Backend Dev", "Non démarré", "Sécurité critique"],
    ["AUTH-3", "Authentification", "Valider inscription", "Dev", "Validation email téléphone et doublons", 4, None, "AUTH-1", "P1", "Moyen", "Backend Dev", "Non démarré", "Contrôles côté service"],
    ["AUTH-4", "Authentification", "Finaliser login", "Dev", "Login avec vérification crédentials et session", 4, None, "AUTH-2", "P1", "Moyen", "Backend Dev", "Non démarré", "Nécessaire pour parcours client"],
    ["AUTH-5", "Authentification", "Tests auth unitaires", "Test", "Tests login inscription erreurs", 6, None, "AUTH-3,AUTH-4", "P1", "Moyen", "QA", "Non démarré", "Inclure cas invalides"],
    ["PROD-1", "Produits", "Concevoir schema produits", "Analysis", "Table produits catégories prix stock", 4, None, "", "P1", "Moyen", "Backend Lead", "Non démarré", "Dépendance catalogue"],
    ["PROD-2", "Produits", "Lister produits avec filtres", "Dev", "Filtrer par catégorie et plage de prix", 6, None, "PROD-1", "P1", "Moyen", "Backend Dev", "Non démarré", "Endpoint principal catalogue"],
    ["PROD-3", "Produits", "Détail produit", "Dev", "Retourner infos détaillées et disponibilité stock", 4, None, "PROD-1", "P2", "Faible", "Backend Dev", "Non démarré", "Vue fiche produit"],
    ["PROD-4", "Produits", "Recherche produits", "Dev", "Recherche textuelle et tri", 8, None, "PROD-2", "P2", "Moyen", "Backend Dev", "Non démarré", "Amélioration UX"],
    ["PROD-5", "Produits", "Tests catalogue", "Test", "Tests filtres recherche et contraintes stock", 6, None, "PROD-2,PROD-3", "P2", "Moyen", "QA", "Non démarré", "Valider cas limites"],
    ["CART-1", "Panier", "Concevoir schema panier", "Analysis", "Tables cart et cart_items", 4, None, "", "P1", "Moyen", "Backend Lead", "Non démarré", "Pré-requis panier"],
    ["CART-2", "Panier", "Ajouter produit panier", "Dev", "Ajout quantité avec contrôle stock", 8, None, "CART-1,PROD-2", "P1", "Élevé", "Backend Dev", "Non démarré", "Point critique métier"],
    ["CART-3", "Panier", "Retirer produit panier", "Dev", "Suppression ligne panier et recalcul", 4, None, "CART-2", "P2", "Faible", "Backend Dev", "Non démarré", "Fonction standard"],
    ["CART-4", "Panier", "Calcul total panier", "Dev", "Sous-total taxes remises éventuelles", 6, None, "CART-2", "P1", "Moyen", "Backend Dev", "Non démarré", "Base checkout"],
    ["CART-5", "Panier", "Persistance panier", "Dev", "Sauvegarde et restauration session panier", 8, None, "CART-1", "P1", "Moyen", "Backend Dev", "Non démarré", "Continuité utilisateur"],
    ["CART-6", "Panier", "Tests panier", "Test", "Tests ajout retrait total et stock", 8, None, "CART-2,CART-4", "P1", "Moyen", "QA", "Non démarré", "Inclure concurrence simple"],
    ["ORDER-1", "Commandes", "Concevoir schema commandes", "Analysis", "Tables orders et order_items", 4, None, "", "P1", "Moyen", "Backend Lead", "Non démarré", "Pré-requis checkout"],
    ["ORDER-2", "Commandes", "Implémenter checkout", "Dev", "Conversion panier vers commande", 12, None, "CART-2,CART-4,ORDER-1", "P1", "Élevé", "Backend Dev", "Non démarré", "Chemin critique"],
    ["ORDER-3", "Commandes", "Workflow statut commande", "Dev", "pending confirmed shipped delivered cancelled", 6, None, "ORDER-2", "P2", "Moyen", "Backend Dev", "Non démarré", "Traçabilité commande"],
    ["ORDER-4", "Commandes", "Historique commandes", "Dev", "Lister commandes client avec filtres date", 6, None, "ORDER-2", "P2", "Faible", "Backend Dev", "Non démarré", "Fonction client"],
    ["ORDER-5", "Commandes", "Tests intégration commandes", "Test", "Parcours E2E checkout statut historique", 12, None, "ORDER-2,ORDER-3", "P1", "Élevé", "QA", "Non démarré", "Valider transactions"],
    ["NET-1", "Réseau Serveur", "Compléter RequestRouter", "Dev", "Gérer tous types de messages métier", 10, None, "AUTH-4,PROD-2,CART-2,ORDER-2", "P0", "Élevé", "Backend Lead", "Non démarré", "Bloque endpoints"],
    ["NET-2", "Réseau Serveur", "Normaliser gestion erreurs", "Dev", "Codes erreurs et réponses JSON homogènes", 6, None, "NET-1", "P1", "Moyen", "Backend Dev", "Non démarré", "Observabilité"],
    ["NET-3", "Réseau Serveur", "Gérer timeouts et cleanup", "Dev", "Fermeture connexions inactives et ressources", 4, None, "NET-1", "P2", "Moyen", "Backend Dev", "Non démarré", "Stabilité serveur"],
    ["NET-4", "Réseau Serveur", "Tests charge concurrence", "Test", "Tester 10+ clients simultanés", 12, None, "NET-1,DB-1", "P1", "Élevé", "QA", "Non démarré", "Mesurer limites"],
    ["DB-1", "Base de données", "Mettre en place pool connexions", "Dev", "Remplacer connexions directes par pool", 8, None, "", "P1", "Moyen", "Backend Dev", "Non démarré", "Performance"],
    ["DB-2", "Base de données", "Script initialisation SQL", "Infra", "Script création schémas index contraintes", 4, None, "AUTH-1,PROD-1,CART-1,ORDER-1", "P1", "Faible", "Backend Dev", "Non démarré", "Onboarding rapide"],
    ["DB-3", "Base de données", "Transactions checkout", "Dev", "Garantir ACID pour opérations commandes stock", 8, None, "ORDER-2", "P0", "Élevé", "Backend Lead", "Non démarré", "Évite incohérences"],
    ["DB-4", "Base de données", "Indexation requêtes clés", "Dev", "Indexer user_id product_id order_id", 4, None, "DB-2", "P2", "Faible", "Backend Dev", "Non démarré", "Optimisation"],
    ["CLIENT-1", "Client", "Améliorer interface console", "Dev", "Menu plus clair et parcours simplifié", 6, None, "AUTH-4,PROD-2", "P3", "Faible", "Frontend Dev", "Non démarré", "Confort utilisateur"],
    ["CLIENT-2", "Client", "Validation saisies client", "Dev", "Vérifier format avant envoi serveur", 4, None, "CLIENT-1", "P2", "Faible", "Frontend Dev", "Non démarré", "Réduction erreurs"],
    ["CLIENT-3", "Client", "Tests client", "Test", "Tests parsing et formatage réponses", 4, None, "CLIENT-1,CLIENT-2", "P3", "Faible", "QA", "Non démarré", "Qualité IHM console"],
    ["DOC-1", "Documentation Déploiement", "Documenter API", "Docs", "Types requêtes payloads réponses exemples", 6, None, "NET-1", "P2", "Faible", "Tech Writer", "Non démarré", "Référence équipe"],
    ["DOC-2", "Documentation Déploiement", "Documenter schéma DB", "Docs", "ERD description tables et relations", 4, None, "DB-2", "P2", "Faible", "Tech Writer", "Non démarré", "Support maintenance"],
    ["DOC-3", "Documentation Déploiement", "Guide déploiement", "Docs", "Pré-requis config lancement server client", 4, None, "DB-2,NET-1", "P2", "Faible", "DevOps", "Non démarré", "Onboarding"],
    ["DEPLOY-1", "Documentation Déploiement", "Configurer build Maven ou Gradle", "Infra", "Automatiser compilation et exécution tests", 6, None, "AUTH-5,PROD-5,CART-6,ORDER-5", "P2", "Moyen", "DevOps", "Non démarré", "Base CI"],
    ["DEPLOY-2", "Documentation Déploiement", "Containeriser application", "Infra", "Dockerfile server client base données", 8, None, "DEPLOY-1,DB-2", "P3", "Moyen", "DevOps", "Non démarré", "Portabilité env"],
    ["DEPLOY-3", "Documentation Déploiement", "Mettre CI CD", "Infra", "Pipeline build tests et packaging", 8, None, "DEPLOY-1", "P2", "Moyen", "DevOps", "Non démarré", "Qualité continue"],
]

wb = Workbook()
ws = wb.active
ws.title = "Backlog_CdC"

for col, header in enumerate(headers, start=1):
    ws.cell(row=1, column=col, value=header)

for r_index, row in enumerate(rows, start=2):
    for c_index, value in enumerate(row, start=1):
        ws.cell(row=r_index, column=c_index, value=value)
    ws.cell(row=r_index, column=7, value=f"=F{r_index}/8")

header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(color="FFFFFF", bold=True)
center_cols = {1, 4, 6, 7, 9, 10, 12}
thin = Side(style="thin", color="D9D9D9")

for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
    for cell in row:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

for row in ws.iter_rows(min_row=2, max_row=1 + len(rows), min_col=1, max_col=len(headers)):
    for cell in row:
        if cell.column in center_cols:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        else:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
        if cell.column == 7:
            cell.number_format = "0.00"

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{1 + len(rows)}"

for col in range(1, len(headers) + 1):
    letter = get_column_letter(col)
    max_len = len(str(ws.cell(row=1, column=col).value))
    for row in range(2, 2 + len(rows)):
        val = ws.cell(row=row, column=col).value
        if val is not None:
            max_len = max(max_len, len(str(val)))
    ws.column_dimensions[letter].width = min(max(12, max_len + 2), 55)

# Summary sheet
sum_ws = wb.create_sheet("Synthese")
sum_ws["A1"] = "KPI"
sum_ws["B1"] = "Valeur"
sum_ws["A2"] = "Nombre total de tâches"
sum_ws["A3"] = "Effort total (h)"
sum_ws["A4"] = "Durée totale (j)"
sum_ws["A5"] = "Tâches critiques (P0/P1)"
sum_ws["A6"] = "Tâches risque élevé"
sum_ws["A7"] = "Tâches terminées"
sum_ws["A8"] = "Avancement global (%)"

sum_ws["B2"] = "=COUNTA(Backlog_CdC!A:A)-1"
sum_ws["B3"] = "=SUM(Backlog_CdC!F:F)"
sum_ws["B4"] = "=SUM(Backlog_CdC!G:G)"
sum_ws["B5"] = "=COUNTIF(Backlog_CdC!I:I,\"P0\")+COUNTIF(Backlog_CdC!I:I,\"P1\")"
sum_ws["B6"] = "=COUNTIF(Backlog_CdC!J:J,\"Élevé\")"
sum_ws["B7"] = "=COUNTIF(Backlog_CdC!L:L,\"Terminé\")"
sum_ws["B8"] = "=IF(B2=0,0,B7/B2)"

sum_ws["A10"] = "Module"
sum_ws["B10"] = "Effort total (h)"
sum_ws["C10"] = "Nombre de tâches"
modules = [
    "Authentification", "Produits", "Panier", "Commandes",
    "Réseau Serveur", "Base de données", "Client", "Documentation Déploiement"
]
for idx, module in enumerate(modules, start=11):
    sum_ws.cell(row=idx, column=1, value=module)
    sum_ws.cell(row=idx, column=2, value=f"=SUMIFS(Backlog_CdC!F:F,Backlog_CdC!B:B,A{idx})")
    sum_ws.cell(row=idx, column=3, value=f"=COUNTIFS(Backlog_CdC!B:B,A{idx})")

sum_header_fill = PatternFill("solid", fgColor="385723")
sum_header_font = Font(color="FFFFFF", bold=True)
for cell in [sum_ws["A1"], sum_ws["B1"], sum_ws["A10"], sum_ws["B10"], sum_ws["C10"]]:
    cell.fill = sum_header_fill
    cell.font = sum_header_font
    cell.alignment = Alignment(horizontal="center")

for row in range(2, 9):
    sum_ws[f"A{row}"].font = Font(bold=True)
    sum_ws[f"B{row}"].alignment = Alignment(horizontal="center")

sum_ws["B8"].number_format = "0.00%"
sum_ws.column_dimensions["A"].width = 40
sum_ws.column_dimensions["B"].width = 20
sum_ws.column_dimensions["C"].width = 20
sum_ws.freeze_panes = "A2"

wb.save(OUTPUT_FILE)
print(f"Generated {OUTPUT_FILE}")
